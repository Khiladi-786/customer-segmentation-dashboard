"""
Report Generator: PDF executive summary and Excel workbook.
"""

from __future__ import annotations
import logging
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

ARTIFACTS_DIR = Path(__file__).parent.parent / "artifacts"
ARTIFACTS_DIR.mkdir(exist_ok=True)


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_excel_report(df: Optional[pd.DataFrame], output_path: str) -> str:
    """Generate a multi-sheet Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        if df is None:
            # Load from models if no df provided
            try:
                import joblib
                df = joblib.load(str(Path(__file__).parent.parent / "models" / "scored_df.pkl"))
            except Exception:
                df = pd.DataFrame({"Note": ["Run Streamlit app first to generate data."]})

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        PURPLE = "7C3AED"
        CYAN   = "06B6D4"

        # ── Sheet 1: Executive Summary ──
        ws_summary = wb.create_sheet("Executive Summary")
        ws_summary["A1"] = "Customer Segmentation Report"
        ws_summary["A1"].font = Font(bold=True, size=16, color=PURPLE)

        if "Cluster" in df.columns:
            summary_data = {
                "Metric":  ["Total Customers", "Segments Found", "Algorithm Used", "Avg CLV Score"],
                "Value":   [
                    len(df),
                    df["Cluster"].nunique(),
                    "AutoML Best",
                    f"{df.get('CLV_Score', pd.Series([0])).mean():.0f}",
                ],
            }
            df_sum = pd.DataFrame(summary_data)
            for r_idx, row in enumerate(dataframe_to_rows(df_sum, index=False, header=True), start=3):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 3:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=PURPLE)

        # ── Sheet 2: Full Data ──
        ws_data = wb.create_sheet("Customer Data")
        cols_to_export = [c for c in df.columns if c not in ("PCA1", "PCA2", "PCA3")]
        df_export = df[cols_to_export].head(2000)
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font  = Font(bold=True, color="FFFFFF")
                    cell.fill  = PatternFill("solid", fgColor=CYAN)

        # ── Sheet per Cluster ──
        if "Cluster" in df.columns:
            for cluster_id in sorted(df["Cluster"].unique()):
                ws = wb.create_sheet(f"Cluster {cluster_id}")
                cluster_df = df[df["Cluster"] == cluster_id][cols_to_export].head(500)
                for r_idx, row in enumerate(dataframe_to_rows(cluster_df, index=False, header=True), start=1):
                    for c_idx, val in enumerate(row, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill("solid", fgColor=PURPLE)

        wb.save(output_path)
        logger.info(f"Excel report saved → {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf_report(output_path: str) -> str:
    """Generate a PDF executive summary report."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import cm

        doc    = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        PURPLE = colors.HexColor("#7c3aed")
        CYAN   = colors.HexColor("#06b6d4")
        DARK   = colors.HexColor("#0a0a0f")

        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            textColor=PURPLE,
            fontSize=24,
            spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            "Heading",
            parent=styles["Heading2"],
            textColor=CYAN,
            fontSize=14,
            spaceAfter=8,
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=11,
            spaceAfter=6,
        )

        story.append(Paragraph("Customer Segmentation — Executive Report", title_style))
        story.append(Paragraph("Generated by AI-Powered Segmentation Platform v2.0", body_style))
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("Platform Overview", heading_style))
        story.append(Paragraph(
            "This report presents the results of an advanced AI-powered customer segmentation analysis using "
            "multiple clustering algorithms with AutoML hyperparameter optimization. The platform identified "
            "optimal customer segments using composite scoring across Silhouette, Davies-Bouldin, and "
            "Calinski-Harabasz metrics.",
            body_style,
        ))
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("Algorithms Evaluated", heading_style))
        algo_data = [
            ["Algorithm", "Description"],
            ["K-Means",            "Centroid-based, fast, interpretable"],
            ["DBSCAN",             "Density-based, handles noise/outliers"],
            ["Agglomerative",      "Hierarchical, no k required upfront"],
            ["Gaussian Mixture",   "Probabilistic, soft cluster assignment"],
        ]
        t = Table(algo_data, colWidths=[6 * cm, 10 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  PURPLE),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f8ff"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("FONTSIZE",    (0, 0), (-1, -1), 10),
            ("PADDING",     (0, 0), (-1, -1), 8),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("Key Features", heading_style))
        features = [
            "✓ RFM Analysis: Recency, Frequency, Monetary segmentation",
            "✓ Customer Lifetime Value (CLV) prediction with GradientBoosting",
            "✓ SHAP Explainability: Why each customer is in their segment",
            "✓ Anomaly Detection: Isolation Forest flags unusual customers",
            "✓ AI Campaign Recommendations via Google Gemini LLM",
            "✓ What-If Simulator: Live cluster prediction from attribute changes",
            "✓ Customer Evolution Tracking: Cluster migration over time",
        ]
        for f in features:
            story.append(Paragraph(f, body_style))
        story.append(Spacer(1, 0.5 * cm))

        doc.build(story)
        logger.info(f"PDF report saved → {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        raise
